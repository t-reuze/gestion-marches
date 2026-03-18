import re
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xlrd
import os
import io
import tkinter as tk
from tkinter import filedialog
from pathlib import Path

# ── Chemins ───────────────────────────────────────────────────────────────────
BASE = Path(r"C:\Users\t-reuze\OneDrive - UNICANCER\Bureau\travail stagiaires _ AO recrutement personnel 2026")
ANNUAIRE_FILE  = BASE / "compilation" / "ANNUAIRE_documents_fournisseurs.xlsx"
COMPIL_FILE    = BASE / "compilation" / "Compilation annexe technique recrutement de personnel.xlsx"
COMPIL_QT_FILE = BASE / "compilation" / "Compilation_QT_recrutement.xlsx"
DCE_QT_FILE    = BASE / "DCE" / "Annexe 1 CCTP  annexe technique recrutement de personnel maj 240226.xls"
REPONSES_DIR   = BASE / "Reponses"

# ── Colonnes de l'annuaire ────────────────────────────────────────────────────
DOC_LABELS = [
    "Lot 1 MAD Personnel",
    "Lot 2 Recrutement",
    "Lot 3 Freelance",
    "BPU (Annexe 5)",
    "Optim. Tarifaire",
    "QT (Annexe 1)",
    "BPU Chiffrage (Annexe 3)",
    "Questionnaire RSE",
    "CCAP signé",
    "CCTP signé",
    "DC1",
    "DC2",
    "ATTRI1",
    "Fiche Contacts",
]
DOC_HEADERS_EXCEL = [
    "Lot 1\nMAD Personnel",
    "Lot 2\nRecrutement",
    "Lot 3\nFreelance",
    "BPU\n(Annexe 5)",
    "Optim.\nTarifaire",
    "QT\n(Annexe 1)",
    "BPU Chiffrage\n(Annexe 3)",
    "Questionnaire\nRSE",
    "CCAP\nsigné",
    "CCTP\nsigné",
    "DC1",
    "DC2",
    "ATTRI1",
    "Fiche\nContacts",
]

# ── Styles Excel — sobre et professionnel ──────────────────────────────────────
FILL_HEADER     = PatternFill("solid", fgColor="1B3A5C")   # bleu marine
FILL_ROW_EVEN   = PatternFill("solid", fgColor="F4F7FA")   # gris très clair
FILL_ROW_ODD    = PatternFill("solid", fgColor="FFFFFF")   # blanc
FILL_MISSING    = PatternFill("solid", fgColor="FDE8E8")   # rose pâle — manquant
FILL_ANNOT      = PatternFill("solid", fgColor="FFF8E1")   # jaune pâle — annotation
FONT_HEADER     = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_NAME       = Font(color="1B3A5C", name="Calibri", size=10)
FONT_X_OK       = Font(bold=True, color="217346", name="Calibri", size=11)
FONT_ANNOT      = Font(italic=True, color="8B5E00", name="Calibri", size=9)
ALIGN_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
THIN_LIGHT      = Side(style="thin",   color="D0D8E4")
THIN_MED        = Side(style="thin",   color="B0BFCE")
BORDER_CELL     = Border(left=THIN_LIGHT, right=THIN_LIGHT, top=THIN_LIGHT, bottom=THIN_LIGHT)
BORDER_NAME     = Border(left=THIN_MED,   right=THIN_LIGHT, top=THIN_LIGHT, bottom=THIN_LIGHT)

# ══════════════════════════════════════════════════════════════════════════════
# Config page — doit être le premier appel Streamlit
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="AO Recrutement Personnel — Traçabilité",
    page_icon="📋",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
# Helpers — Détection automatique
# ══════════════════════════════════════════════════════════════════════════════

def sheet_has_data(sh, min_values: int = 3) -> bool:
    count = 0
    try:
        if hasattr(sh, "row_values"):
            for r in range(sh.nrows):
                for v in sh.row_values(r):
                    if isinstance(v, float) and v != 0.0:
                        count += 1
                        if count >= min_values:
                            return True
        else:
            for row in sh.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, (int, float)) and v != 0:
                        count += 1
                        if count >= min_values:
                            return True
    except Exception:
        pass
    return False

SHEET_TO_LOT = {
    "lot 1": 1, "lot 1 mad": 1, "personnel": 1,
    "lot 2": 2, "lot 2 rec": 2, "recrutement": 2,
    "lot 3": 3, "lot 3 fre": 3, "freelance": 3,
}

def sheet_lot(sheet_name: str):
    n = sheet_name.lower()
    if "optim" in n:
        return "optim"
    for key, lot in SHEET_TO_LOT.items():
        if key in n:
            return lot
    return None

def find_annexe3_files(fournisseur_dir: Path) -> list:
    results = []
    for f in fournisseur_dir.rglob("*"):
        if not f.is_file() or f.name.startswith("~"):
            continue
        if f.suffix.lower() not in (".xls", ".xlsx", ".pdf", ".p7m"):
            continue
        n = f.name.lower()
        if "annexe 3" in n or "annexe_3" in n or "chiffrage" in n:
            results.append(f)
    return results

def detect_lots_from_annexe3(annexe3_files: list) -> set:
    covered = set()
    for f in annexe3_files:
        if f.suffix.lower() in (".pdf", ".p7m"):
            found = lot_from_filename(f.name)
            if not found:
                for parent in f.parents:
                    found = lot_from_filename(parent.name)
                    if found:
                        break
            covered.update(found)
            continue
        filename_lots = lot_from_filename(f.name)
        if not filename_lots:
            for parent in f.parents:
                pl = lot_from_filename(parent.name)
                if pl:
                    filename_lots = pl
                    break
        if filename_lots:
            covered.update(filename_lots)
            continue
        try:
            if f.suffix.lower() == ".xls":
                wb = xlrd.open_workbook(str(f))
                sheets = {sh.name: sh for sh in wb.sheets()}
            else:
                wb = openpyxl.load_workbook(str(f), data_only=True, read_only=True)
                sheets = {name: wb[name] for name in wb.sheetnames}
            for name, sh in sheets.items():
                lot = sheet_lot(name)
                if isinstance(lot, int) and sheet_has_data(sh):
                    covered.add(lot)
        except Exception:
            for lot in lot_from_filename(f.name):
                covered.add(lot)
    return covered

def find_annexe5_files(fournisseur_dir: Path) -> list:
    EXCL = ("annexe 3", "annexe_3", "chiffrage")
    results = []
    for f in fournisseur_dir.rglob("*"):
        if not f.is_file() or f.name.startswith("~"):
            continue
        if f.suffix.lower() not in (".xls", ".xlsx", ".pdf", ".p7m"):
            continue
        n = f.name.lower()
        if any(e in n for e in EXCL):
            continue
        if ("annexe 5" in n or "annexe_5" in n or "bpu" in n
                or "bordereau de prix" in n or "bordereau_de_prix" in n):
            results.append(f)
    return results

def lot_from_filename(filename: str) -> list:
    n = filename.lower()
    lots = set()
    for m in re.finditer(r'lot[\s_-]*(\d[\d\s,/&etET+-]*)', n):
        for d in re.findall(r'[123]', m.group(0)):
            lots.add(int(d))
    return sorted(lots)

def detect_lots_from_bpu(bpu_files: list) -> dict:
    lots = {1: False, 2: False, 3: False, "optim": False}
    has_bpu = len(bpu_files) > 0

    for f in bpu_files:
        if f.suffix.lower() in (".pdf", ".p7m"):
            found = lot_from_filename(f.name)
            if not found:
                for parent in f.parents:
                    found = lot_from_filename(parent.name)
                    if found:
                        break
            for lot in found:
                lots[lot] = True
            has_bpu = True
            continue

        filename_lots = lot_from_filename(f.name)
        if not filename_lots:
            for parent in f.parents:
                parent_lots = lot_from_filename(parent.name)
                if parent_lots:
                    filename_lots = parent_lots
                    break
        if filename_lots:
            for lot in filename_lots:
                lots[lot] = True
            try:
                if f.suffix.lower() == ".xls":
                    wb = xlrd.open_workbook(str(f))
                    sheets = {sh.name: sh for sh in wb.sheets()}
                else:
                    wb = openpyxl.load_workbook(str(f), data_only=True, read_only=True)
                    sheets = {name: wb[name] for name in wb.sheetnames}
                for name, sh in sheets.items():
                    if sheet_lot(name) == "optim" and sheet_has_data(sh, min_values=5):
                        lots["optim"] = True
            except Exception:
                pass
            continue

        try:
            if f.suffix.lower() == ".xls":
                wb = xlrd.open_workbook(str(f))
                sheets = {sh.name: sh for sh in wb.sheets()}
            else:
                wb = openpyxl.load_workbook(str(f), data_only=True, read_only=True)
                sheets = {name: wb[name] for name in wb.sheetnames}

            found_any_sheet = False
            for name, sh in sheets.items():
                lot = sheet_lot(name)
                if lot is None:
                    continue
                threshold = 5 if lot == "optim" else 3
                if sheet_has_data(sh, min_values=threshold):
                    lots[lot] = True
                    found_any_sheet = True

            if not found_any_sheet:
                for lot in lot_from_filename(f.name):
                    lots[lot] = True

        except Exception:
            for lot in lot_from_filename(f.name):
                lots[lot] = True

    return {**{f"lot{k}": v for k, v in lots.items() if k != "optim"},
            "optim": lots["optim"],
            "has_bpu": has_bpu}

def detect_supplier(fournisseur_dir: Path) -> dict:
    all_lower = [f.name.lower() for f in fournisseur_dir.rglob("*")
                 if f.is_file() and not f.name.startswith("~")]

    val = lambda b: "x" if b else ""

    bpu_files = find_annexe5_files(fournisseur_dir)
    lots_info = detect_lots_from_bpu(bpu_files)

    annexe3_files = find_annexe3_files(fournisseur_dir)
    if annexe3_files:
        lots_fournisseur = {k for k in [1, 2, 3] if lots_info.get(f"lot{k}")}
        lots_annexe3     = detect_lots_from_annexe3(annexe3_files)
        chiffrage = bool(lots_fournisseur) and lots_fournisseur.issubset(lots_annexe3)
    else:
        chiffrage = False

    qt = any(
        ("qt_lot" in n or "qt lot" in n or ("annexe" in n and "1" in n and "cctp" in n))
        and (n.endswith((".xls", ".xlsx")) or n.endswith((".xls.p7m", ".xlsx.p7m")))
        for n in all_lower
    )
    rse  = any("rse" in n for n in all_lower)
    ccap = any("ccap" in n and "bpu" not in n and "annexe 5" not in n
                and (n.endswith(".pdf") or n.endswith(".p7m")) for n in all_lower)
    cctp = any("cctp" in n and "annexe 1" not in n and "annexe_1" not in n
                and "qt" not in n
                and (n.endswith(".pdf") or n.endswith(".p7m")) for n in all_lower)
    dc1      = any(n.startswith("dc1") for n in all_lower)
    dc2      = any(n.startswith("dc2") for n in all_lower)
    attri    = any("attri1" in n or ("attri" in n and "sign" in n) for n in all_lower)
    contacts = any("contact" in n or "annexe 4" in n or "annexe_4" in n for n in all_lower)

    if lots_info["has_bpu"] and not any([lots_info["lot1"], lots_info["lot2"], lots_info["lot3"]]):
        for n in all_lower:
            for lot in lot_from_filename(n):
                lots_info[f"lot{lot}"] = True

    return {
        "Lot 1 MAD Personnel":      val(lots_info["lot1"]),
        "Lot 2 Recrutement":        val(lots_info["lot2"]),
        "Lot 3 Freelance":          val(lots_info["lot3"]),
        "BPU (Annexe 5)":           val(lots_info["has_bpu"]),
        "Optim. Tarifaire":         val(lots_info["optim"]),
        "QT (Annexe 1)":            val(qt),
        "BPU Chiffrage (Annexe 3)": val(chiffrage),
        "Questionnaire RSE":        val(rse),
        "CCAP signé":               val(ccap),
        "CCTP signé":               val(cctp),
        "DC1":                      val(dc1),
        "DC2":                      val(dc2),
        "ATTRI1":                   val(attri),
        "Fiche Contacts":           val(contacts),
    }

def scan_all_suppliers_in(reponses_dir: Path) -> list[dict]:
    if not reponses_dir.exists():
        return []
    results = []
    for d in sorted(reponses_dir.iterdir()):
        if not d.is_dir():
            continue
        row = {"Nom fournisseur": d.name}
        row.update(detect_supplier(d))
        results.append(row)
    return results

# ── Helpers — Lecture / écriture annuaire Excel ───────────────────────────────

def load_annuaire_excel() -> list[dict]:
    wb = openpyxl.load_workbook(ANNUAIRE_FILE)
    ws = wb["ANNUAIRE"]
    rows = list(ws.iter_rows(values_only=True))
    result = []
    for row in rows[1:]:
        if not row[0]:
            continue
        d = {"Nom fournisseur": row[0]}
        for i, label in enumerate(DOC_LABELS):
            raw = row[i + 1] if (i + 1) < len(row) else None
            d[label] = str(raw).strip() if raw is not None else ""
        result.append(d)
    return result

def build_annuaire_excel(df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ANNUAIRE"

    headers = ["Nom fournisseur"] + DOC_HEADERS_EXCEL

    # ── Ligne d'en-tête ───────────────────────────────────────────────────────
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill      = FILL_HEADER
        cell.font      = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_CELL
    ws.row_dimensions[1].height = 54

    # ── Largeurs colonnes ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    col_widths = {
        2: 10, 3: 11, 4: 10,
        5: 11, 6: 11,
        7: 10, 8: 13,
        9: 13,
        10: 9, 11: 9,
        12: 8, 13: 8,
        14: 9, 15: 11,
    }
    for col_idx in range(2, len(headers) + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_idx, 9)

    # ── Données ───────────────────────────────────────────────────────────────
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        doc_vals   = [str(row.get(lbl, "")).strip() for lbl in DOC_LABELS]
        row_fill   = FILL_ROW_EVEN if r_idx % 2 == 0 else FILL_ROW_ODD

        nom = str(row.get("Nom fournisseur", "")).strip()
        nom_clean = re.sub(r'\s+ok\s*$', '', nom, flags=re.IGNORECASE)
        cell_a           = ws.cell(row=r_idx, column=1, value=nom_clean)
        cell_a.fill      = row_fill
        cell_a.font      = FONT_NAME
        cell_a.alignment = ALIGN_LEFT
        cell_a.border    = BORDER_NAME

        for c_idx, label in enumerate(DOC_LABELS, 2):
            val  = str(row.get(label, "")).strip()
            cell = ws.cell(row=r_idx, column=c_idx, value=None)
            cell.alignment = ALIGN_CENTER
            cell.border    = BORDER_CELL

            if val.lower() == "x":
                cell.value = "●"
                cell.font  = FONT_X_OK
                cell.fill  = row_fill
            elif val:
                cell.value = val
                cell.font  = FONT_ANNOT
                cell.fill  = FILL_ANNOT
            else:
                cell.fill = FILL_MISSING

        ws.row_dimensions[r_idx].height = 20

    ws.freeze_panes = "B2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def save_annuaire_excel(df: pd.DataFrame):
    excel_bytes = build_annuaire_excel(df)
    wb_orig = openpyxl.load_workbook(ANNUAIRE_FILE)
    if "ANNUAIRE" in wb_orig.sheetnames:
        del wb_orig["ANNUAIRE"]
    wb_new  = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws_new  = wb_new.active
    ws_copy = wb_orig.create_sheet("ANNUAIRE", 0)
    for row in ws_new.iter_rows():
        for cell in row:
            new_cell = ws_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font      = cell.font.copy()
                new_cell.fill      = cell.fill.copy()
                new_cell.border    = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
    for k, v in ws_new.column_dimensions.items():
        ws_copy.column_dimensions[k].width = v.width
    for k, v in ws_new.row_dimensions.items():
        ws_copy.row_dimensions[k].height = v.height
    ws_copy.freeze_panes = "B2"
    wb_orig.save(ANNUAIRE_FILE)

def load_contacts_excel():
    wb = openpyxl.load_workbook(ANNUAIRE_FILE)
    ws = wb["CONTACTS"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data = [r for r in rows[1:] if any(v is not None for v in r)]
    return headers, data

def save_contacts_excel(headers, data):
    wb = openpyxl.load_workbook(ANNUAIRE_FILE)
    ws = wb["CONTACTS"]
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(ANNUAIRE_FILE)

# ── Helpers — Lecture QT ──────────────────────────────────────────────────────

def read_qt_file(path: Path) -> list:
    rows = []
    try:
        if path.suffix.lower() == ".xls":
            wb = xlrd.open_workbook(str(path))
            sh = wb.sheets()[0]
            for r in range(sh.nrows):
                rv  = sh.row_values(r)
                q   = str(rv[0]).strip() if len(rv) > 0 and rv[0] else ""
                det = str(rv[1]).strip() if len(rv) > 1 and rv[1] else ""
                rep = str(rv[2]).strip() if len(rv) > 2 and rv[2] else ""
                rows.append((q, det, rep))
        else:
            wb = openpyxl.load_workbook(str(path), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                q   = str(row[0]).strip() if row[0] is not None else ""
                det = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                rep = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                rows.append((q, det, rep))
    except Exception as e:
        st.warning(f"Impossible de lire {path.name} : {e}")
    return rows

def read_template_qt(lot: int) -> list:
    try:
        wb = xlrd.open_workbook(str(DCE_QT_FILE))
    except Exception as e:
        st.error(f"Impossible de lire le template DCE : {e}")
        return []
    sheet_name = f"QT LOT {lot}"
    try:
        sh = wb.sheet_by_name(sheet_name)
    except Exception:
        return []
    questions = []
    for r in range(sh.nrows):
        row = sh.row_values(r)
        q = str(row[0]).strip() if row[0] else ""
        if not q:
            continue
        questions.append((r, q))
    return questions

def get_supplier_qt(fournisseur_dir: Path, lot: int):
    tech_dir = None
    for d in fournisseur_dir.iterdir():
        if d.is_dir() and ("organisationnel" in d.name.lower() or "technique" in d.name.lower()):
            tech_dir = d
            break
    search_dir = tech_dir if tech_dir else fournisseur_dir

    EXCLUDE = ("annexe 3", "annexe_3", "annexe 5", "annexe_5", "bpu", "attri", "chiffrage", "rse")

    candidates = []
    for f in search_dir.glob("*"):
        if not f.is_file() or f.name.startswith("~"):
            continue
        if f.suffix.lower() not in (".xls", ".xlsx"):
            continue
        n = f.name.lower()
        if any(excl in n for excl in EXCLUDE):
            continue
        lot_match = (f"lot_{lot}" in n or f"lot {lot}" in n or f"lot{lot}" in n)
        is_qt = "qt" in n or ("annexe" in n and "1" in n and "cctp" in n)
        if lot_match and is_qt:
            candidates.append(f)

    if not candidates:
        for f in search_dir.glob("*"):
            if not f.is_file() or f.name.startswith("~"):
                continue
            if f.suffix.lower() not in (".xls", ".xlsx"):
                continue
            n = f.name.lower()
            if any(excl in n for excl in EXCLUDE):
                continue
            if f"qt_lot_{lot}" in n or f"qt_lot{lot}" in n:
                candidates.append(f)

    if not candidates:
        return None
    for c in candidates:
        if f"lot_{lot}" in c.name.lower() or f"lot {lot}" in c.name.lower():
            return c
    return candidates[0]

# ── Compilation QT ────────────────────────────────────────────────────────────

def compile_qt_lot(lot: int, suppliers: list) -> tuple:
    template = read_template_qt(lot)
    if not template:
        return pd.DataFrame(), {}

    rows_data = [{"_row": idx, "Question": question} for idx, question in template]
    sup_status = {}
    template_rows = {idx for idx, _ in template}

    for sup_dir in suppliers:
        qt_file  = get_supplier_qt(sup_dir, lot)
        sup_name = sup_dir.name.replace(" ok", "").replace(" OK", "").strip().upper()
        answers_by_row = {}
        if qt_file:
            for r, (q, det, rep) in enumerate(read_qt_file(qt_file)):
                answers_by_row[r] = rep
            filled = sum(1 for idx in template_rows if answers_by_row.get(idx, "").strip())
            total  = len(template_rows)
            sup_status[sup_name] = "ok" if filled == total else "partial" if filled > 0 else "empty"
        else:
            sup_status[sup_name] = "absent"

        for rowinfo in rows_data:
            r = rowinfo["_row"]
            rowinfo[sup_name] = answers_by_row.get(r, "")

    df = pd.DataFrame(rows_data).drop(columns=["_row"])
    return df, sup_status

def write_compilation_xlsx(dfs_and_status: dict) -> bytes:
    HDR_FILL    = PatternFill("solid", fgColor="1F3864")
    HDR_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    Q_FILL      = PatternFill("solid", fgColor="D9E1F2")
    Q_FONT      = Font(name="Calibri", size=10)
    OK_FILL     = PatternFill("solid", fgColor="E2EFDA")
    MISS_FILL   = PatternFill("solid", fgColor="FCE4D6")
    ABSENT_FILL = PatternFill("solid", fgColor="F4CCCC")
    WRAP_TOP    = Alignment(wrap_text=True, vertical="top")
    WRAP_CTR    = Alignment(wrap_text=True, vertical="center", horizontal="center")

    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)

    for lot, (df, sup_status) in dfs_and_status.items():
        if df.empty:
            continue
        ws       = wb.create_sheet(title=f"QT LOT {lot}")
        sup_cols = [c for c in df.columns if c != "Question"]

        ws.cell(row=1, column=1, value="Question").fill      = HDR_FILL
        ws.cell(row=1, column=1).font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        ws.cell(row=1, column=1).alignment = WRAP_CTR
        for c_idx, sup in enumerate(sup_cols, 2):
            cell           = ws.cell(row=1, column=c_idx, value=sup)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = WRAP_CTR

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            q_cell           = ws.cell(row=r_idx, column=1, value=row["Question"])
            q_cell.fill      = Q_FILL
            q_cell.font      = Q_FONT
            q_cell.alignment = WRAP_TOP
            for c_idx, sup in enumerate(sup_cols, 2):
                val    = str(row.get(sup, "")).strip()
                cell   = ws.cell(row=r_idx, column=c_idx, value=val if val else None)
                cell.alignment = WRAP_TOP
                status = sup_status.get(sup, "absent")
                if status == "absent":
                    cell.fill = ABSENT_FILL
                elif val:
                    cell.fill = OK_FILL
                else:
                    cell.fill = MISS_FILL

        ws.column_dimensions["A"].width = 55
        for c_idx in range(2, len(sup_cols) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 38
        ws.freeze_panes = "B2"

    if not wb.worksheets:
        ws = wb.create_sheet(title="Vide")
        ws.cell(row=1, column=1, value="Aucune donnée à compiler.")

    buf.seek(0)
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def qt_detail_status(fournisseur_dir: Path, lot: int, template: list) -> dict:
    qt_file = get_supplier_qt(fournisseur_dir, lot)
    total   = len(template)
    if not qt_file:
        return {"fichier": "—", "statut": "❌ Fichier absent", "remplies": 0, "total": total}
    template_rows = {idx for idx, _ in template}
    rows   = read_qt_file(qt_file)
    filled = sum(1 for r, (q, det, rep) in enumerate(rows)
                 if r in template_rows and rep.strip())
    if filled == total:
        statut = f"✅ Complet ({filled}/{total})"
    elif filled > 0:
        statut = f"⚠️ Partiel ({filled}/{total} — {total - filled} manquantes)"
    else:
        statut = f"❌ Fichier présent mais vide (0/{total})"
    return {"fichier": qt_file.name, "statut": statut, "remplies": filled, "total": total}


# ══════════════════════════════════════════════════════════════════════════════
# CSS personnalisé
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Base ── */
.block-container { padding-top: 2.5rem !important; max-width: 1400px; }

/* ── Bandeau UNICANCER ── */
.unicancer-banner {
    background: #1B3A5C;
    border-radius: 8px;
    padding: 1rem 1.6rem;
    margin-top: 0.2rem;
    margin-bottom: 1.2rem;
}
.unicancer-banner .uc-tag {
    display: inline-block;
    background: #E87722;
    color: #fff;
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 4px;
    margin-bottom: 0.45rem;
}
.unicancer-banner .uc-title {
    color: #ffffff;
    font-size: 1.25rem;
    font-weight: 700;
    margin: 0 0 2px;
}
.unicancer-banner .uc-sub {
    color: rgba(255,255,255,0.65);
    font-size: 0.82rem;
    margin: 0;
    border-top: 1px solid rgba(232,119,34,0.4);
    padding-top: 0.4rem;
    margin-top: 0.4rem;
}

/* ── Métriques ── */
div[data-testid="stMetric"] {
    background: #ffffff;
    border-radius: 8px;
    padding: 12px 16px;
    border: 1px solid #E0E7F0;
    border-left: 4px solid #E87722;
}
div[data-testid="stMetricLabel"] p {
    color: #5a6a7a !important;
    font-size: .82rem !important;
}
div[data-testid="stMetricValue"] {
    color: #1B3A5C !important;
    font-size: 1.7rem !important;
    font-weight: 700 !important;
}

/* ── Boutons primaires orange ── */
button[kind="primary"] {
    background: #E87722 !important;
    border-color: #E87722 !important;
    color: #fff !important;
    font-weight: 600 !important;
}
button[kind="primary"]:hover {
    background: #C96010 !important;
    border-color: #C96010 !important;
}

/* ── Onglets ── */
button[data-baseweb="tab"] { color: #1B3A5C !important; font-weight: 600 !important; }
button[data-baseweb="tab"][aria-selected="true"] {
    color: #E87722 !important;
    border-bottom: 3px solid #E87722 !important;
}

/* ── Cartes ── */
div[data-testid="stVerticalBlockBorderWrapper"] {
    border-radius: 8px !important;
    border-color: #D5E0EE !important;
}
div[data-testid="stVerticalBlockBorderWrapper"] h3 {
    color: #1B3A5C !important;
    font-weight: 700 !important;
}

/* ── Divider ── */
hr { border-color: #D5E0EE !important; opacity: 1; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# En-tête
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="unicancer-banner">
    <span class="uc-tag">UNICANCER</span>
    <p class="uc-title">AO Recrutement de Personnel 2026</p>
    <p class="uc-sub">Traçabilité des documents fournisseurs &nbsp;&middot;&nbsp; Compilation des Questionnaires Techniques</p>
</div>
""", unsafe_allow_html=True)

# ── Métriques rapides ─────────────────────────────────────────────────────────
if "annuaire_df" in st.session_state and not st.session_state["annuaire_df"].empty:
    df_m    = st.session_state["annuaire_df"]
    nb_f    = len(df_m)
    doc_cols_m = [c for c in DOC_LABELS if c in df_m.columns]
    m1, m2 = st.columns(2)
    m1.metric("Fournisseurs analysés", nb_f)
    m2.metric("QT compilés", "Oui" if "qt_excel" in st.session_state else "Non")

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# ZONE D'ACTIONS — 3 cartes
# ══════════════════════════════════════════════════════════════════════════════
col_a, col_b, col_c = st.columns(3, gap="medium")

# ── Carte 1 : Dossier source ──────────────────────────────────────────────────
with col_a:
    with st.container(border=True):
        st.markdown("### 📁 &nbsp;Dossier source")
        reponses_dir = Path(st.session_state.get("reponses_dir", str(REPONSES_DIR)))
        st.caption("Dossier actuel :")
        st.code(str(reponses_dir), language=None)
        if st.button("📂  Parcourir…", use_container_width=True, key="browse_folder"):
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes('-topmost', 1)
            folder = filedialog.askdirectory(
                title="Sélectionner le dossier Réponses",
                initialdir=str(reponses_dir) if reponses_dir.exists() else str(Path.home()),
            )
            root.destroy()
            if folder:
                st.session_state["reponses_dir"] = folder
                st.session_state.pop("annuaire_df", None)
                st.session_state.pop("qt_excel",    None)
                st.rerun()

# ── Carte 2 : Annuaire documents ─────────────────────────────────────────────
with col_b:
    with st.container(border=True):
        st.markdown("### 📊 &nbsp;Annuaire documents")

        reponses_dir2 = Path(st.session_state.get("reponses_dir", str(REPONSES_DIR)))

        if st.button("🔍  Analyser le dossier", type="primary", use_container_width=True):
            with st.spinner("Scan en cours…"):
                scanned = scan_all_suppliers_in(reponses_dir2)
            if scanned:
                st.session_state["annuaire_df"] = pd.DataFrame(scanned)
                st.success(f"✅ {len(scanned)} fournisseurs détectés")
                st.rerun()
            else:
                st.error("Aucun dossier fournisseur trouvé dans ce répertoire.")

        st.markdown("")

        annuaire_ready = "annuaire_df" in st.session_state and not st.session_state["annuaire_df"].empty

        if annuaire_ready:
            excel_ann = build_annuaire_excel(st.session_state["annuaire_df"])
            st.download_button(
                label="📥  Exporter l'annuaire Excel",
                data=excel_ann,
                file_name="ANNUAIRE_documents_fournisseurs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            if st.button("💾  Sauvegarder l'annuaire", use_container_width=True):
                try:
                    save_annuaire_excel(st.session_state["annuaire_df"])
                    st.success("Annuaire sauvegardé ✅")
                except Exception as e:
                    st.error(f"Erreur : {e}")
        else:
            st.button("📥  Exporter l'annuaire Excel", disabled=True, use_container_width=True,
                      help="Lancez d'abord l'analyse du dossier.")
            st.button("💾  Sauvegarder l'annuaire",   disabled=True, use_container_width=True)

# ── Carte 3 : Compilation QT ──────────────────────────────────────────────────
with col_c:
    with st.container(border=True):
        st.markdown("### 📋 &nbsp;Compilation QT")

        reponses_dir3  = Path(st.session_state.get("reponses_dir", str(REPONSES_DIR)))
        suppliers_dirs = sorted([d for d in reponses_dir3.iterdir() if d.is_dir()]) \
            if reponses_dir3.exists() else []

        if suppliers_dirs:
            lots_sel = st.multiselect(
                "Lots à compiler",
                options=[1, 2, 3],
                default=[1, 2, 3],
                key="lots_sel_card",
                format_func=lambda x: f"LOT {x}",
            )
            if st.button("⚙️  Compiler les QT", type="primary", use_container_width=True,
                         disabled=not lots_sel):
                dfs_and_status = {}
                prog = st.progress(0, text="Compilation en cours…")
                for i, lot in enumerate(lots_sel):
                    prog.progress((i) / len(lots_sel), text=f"Compilation LOT {lot}…")
                    dfs_and_status[lot] = compile_qt_lot(lot, suppliers_dirs)
                prog.progress(1.0, text="Terminé ✅")
                qt_excel_bytes = write_compilation_xlsx(dfs_and_status)
                st.session_state["qt_excel"] = qt_excel_bytes
                try:
                    with open(COMPIL_QT_FILE, "wb") as f:
                        f.write(qt_excel_bytes)
                except Exception:
                    pass
                st.success("Compilation terminée ✅")
                st.rerun()
        else:
            st.caption("Aucun fournisseur trouvé. Vérifiez le dossier source.")
            st.button("⚙️  Compiler les QT", disabled=True, use_container_width=True)

        st.markdown("")

        if "qt_excel" in st.session_state:
            st.download_button(
                label="📥  Exporter la compilation Excel",
                data=st.session_state["qt_excel"],
                file_name="Compilation_QT_recrutement.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.button("📥  Exporter la compilation Excel", disabled=True, use_container_width=True,
                      help="Lancez d'abord la compilation des QT.")

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLETS DÉTAIL
# ══════════════════════════════════════════════════════════════════════════════
tab1, tab2, tab3 = st.tabs(["📊  Annuaire documents", "👥  Contacts fournisseurs", "📋  Détail QT"])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — ANNUAIRE
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    st.caption("Tapez **x** si le document est présent, laissez vide sinon. "
               "Une note (ex : *Illisible*, *Pas de QT*) apparaît en orange dans l'Excel exporté.")

    if "annuaire_df" not in st.session_state:
        try:
            rows = load_annuaire_excel()
            st.session_state["annuaire_df"] = pd.DataFrame(rows) if rows else \
                pd.DataFrame(columns=["Nom fournisseur"] + DOC_LABELS)
        except Exception:
            st.session_state["annuaire_df"] = pd.DataFrame(columns=["Nom fournisseur"] + DOC_LABELS)

    df_display = st.session_state["annuaire_df"]

    edited = st.data_editor(
        df_display,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "Nom fournisseur": st.column_config.TextColumn("Fournisseur", width="medium"),
            **{col: st.column_config.TextColumn(col, width="small", max_chars=30)
               for col in DOC_LABELS if col in df_display.columns},
        },
    )
    st.session_state["annuaire_df"] = edited

    st.divider()
    st.markdown("**Récapitulatif — documents manquants par fournisseur**")
    recap = []
    doc_cols_present = [c for c in DOC_LABELS if c in edited.columns]
    for _, row in edited.iterrows():
        present   = [c for c in doc_cols_present if str(row.get(c, "")).strip()]
        manquants = [c for c in doc_cols_present if not str(row.get(c, "")).strip()]
        recap.append({
            "Fournisseur":  row["Nom fournisseur"],
            "Reçus":        f"{len(present)}/{len(doc_cols_present)}",
            "Manquants":    ", ".join(manquants) if manquants else "—",
        })
    st.dataframe(pd.DataFrame(recap), use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — CONTACTS
# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    st.subheader("Contacts fournisseurs")
    try:
        c_headers, c_data = load_contacts_excel()
        contacts_df = pd.DataFrame(c_data, columns=c_headers)
        edited_contacts = st.data_editor(
            contacts_df,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            column_config={
                "Nom fournisseur": st.column_config.TextColumn(width="medium"),
                "Prénom":          st.column_config.TextColumn(width="small"),
                "Nom":             st.column_config.TextColumn(width="small"),
                "Téléphone":       st.column_config.TextColumn(width="medium"),
                "Email":           st.column_config.TextColumn(width="large"),
            },
        )
        if st.button("💾  Sauvegarder les contacts", type="primary"):
            save_contacts_excel(c_headers, [tuple(r) for _, r in edited_contacts.iterrows()])
            st.success("Contacts sauvegardés ✅")
    except Exception as e:
        st.warning(f"Impossible de charger la feuille CONTACTS : {e}")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — DÉTAIL QT
# ─────────────────────────────────────────────────────────────────────────────
with tab3:
    st.subheader("Détail des QT par fournisseur et par lot")

    reponses_dir_qt   = Path(st.session_state.get("reponses_dir", str(REPONSES_DIR)))
    suppliers_dirs_qt = sorted([d for d in reponses_dir_qt.iterdir() if d.is_dir()]) \
        if reponses_dir_qt.exists() else []

    if not suppliers_dirs_qt:
        st.warning("Aucun dossier fournisseur trouvé. Vérifiez le dossier source.")
    else:
        sup_names_qt    = [d.name for d in suppliers_dirs_qt]
        selected_sups_qt = st.multiselect(
            "Fournisseurs à afficher",
            options=sup_names_qt,
            default=sup_names_qt,
            key="qt_detail_sups",
        )

        templates_cache = {}
        qt_status_rows  = []
        for d in suppliers_dirs_qt:
            if d.name not in selected_sups_qt:
                continue
            for lot in [1, 2, 3]:
                if lot not in templates_cache:
                    templates_cache[lot] = read_template_qt(lot)
                tmpl   = templates_cache[lot]
                detail = qt_detail_status(d, lot, tmpl)
                qt_status_rows.append({
                    "Fournisseur": d.name,
                    "Lot":         f"LOT {lot}",
                    "Fichier":     detail["fichier"],
                    "Statut":      detail["statut"],
                    "Remplies":    f"{detail['remplies']}/{detail['total']}",
                })

        st.dataframe(pd.DataFrame(qt_status_rows), use_container_width=True, hide_index=True)
